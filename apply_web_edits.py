#!/usr/bin/env python3
"""Pull the repo and apply web edits (rows flagged _edited in data.json) back into Job_Application_Tracker.xlsx.

Web-editable fields: Status + Notes (Applications), Status + Fit Notes (New Jobs).
Rows are matched by hyperlink URL first, then by Company+Role. Run before rebuilding the site.

After edits are applied, any "New Jobs" row whose Status was moved past "Saved"/"In progress"
(i.e. Applied, Submitted, Interview, Offer, Rejected, Not proceeding) is migrated into the
Applications table, since that reflects a job actually applied to rather than just queued.
"""
import json
import os
import subprocess
from datetime import date
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Font

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "Job_Application_Tracker.xlsx")

NON_MIGRATING_STATUSES = {"Saved", "In progress", None, ""}

subprocess.run(["git", "-C", HERE, "pull", "--ff-only"], check=False)

with open(os.path.join(HERE, "data.json")) as f:
    data = json.load(f)

SHEETS = {
    # sheet name: (json key, link col header, editable {json field: column header})
    "Applications": ("applications", "Job Link", {"Status": "Status", "Notes": "Notes"}),
    "New Jobs": ("new_jobs", "Job Link", {"Status": "Status", "Fit Notes": "Fit Notes"}),
}

wb = openpyxl.load_workbook(XLSX)
applied = 0

for sheet_name, (key, link_col, editable) in SHEETS.items():
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    headers = {c.value: i for i, c in enumerate(ws[1], start=1)}
    edited = [r for r in data.get(key, []) if r.get("_edited")]
    if not edited:
        continue
    for rec in edited:
        for row in ws.iter_rows(min_row=2):
            link_cell = row[headers[link_col] - 1] if link_col in headers else None
            url = link_cell.hyperlink.target if (link_cell and link_cell.hyperlink) else None
            same_url = rec.get("_url") and url == rec["_url"]
            same_key = (
                str(row[headers["Company"] - 1].value) == str(rec.get("Company"))
                and str(row[headers["Role"] - 1].value) == str(rec.get("Role"))
            )
            if same_url or same_key:
                for field, col in editable.items():
                    if col in headers and rec.get(field) is not None:
                        ws.cell(row=row[0].row, column=headers[col], value=str(rec[field]).strip())
                applied += 1
                break

migrated = 0
if "New Jobs" in wb.sheetnames and "Applications" in wb.sheetnames:
    nj = wb["New Jobs"]
    ap = wb["Applications"]
    nj_headers = {c.value: i for i, c in enumerate(nj[1], start=1)}
    ap_headers = {c.value: i for i, c in enumerate(ap[1], start=1)}
    n_cols = len(nj_headers)

    # Snapshot every data row's values + hyperlink target before touching the sheet,
    # since openpyxl's delete_rows does not reliably relocate Hyperlink objects.
    all_rows = []
    for row in nj.iter_rows(min_row=2):
        if all(c.value in (None, "") for c in row):
            continue
        values = [c.value for c in row]
        link_cell = row[nj_headers["Job Link"] - 1]
        url = link_cell.hyperlink.target if link_cell.hyperlink else None
        all_rows.append({"values": values, "url": url})

    keep_rows = []
    migrate_rows = []
    for r in all_rows:
        status = r["values"][nj_headers["Status"] - 1]
        (migrate_rows if status not in NON_MIGRATING_STATUSES else keep_rows).append(r)

    for r in migrate_rows:
        vals = r["values"]
        company = vals[nj_headers["Company"] - 1]
        role = vals[nj_headers["Role"] - 1]
        location = vals[nj_headers["Location"] - 1]
        status = vals[nj_headers["Status"] - 1]
        fit_notes = vals[nj_headers["Fit Notes"] - 1]
        url = r["url"]

        ap_row = ap.max_row + 1
        ap.cell(ap_row, ap_headers["Company"], company)
        ap.cell(ap_row, ap_headers["Role"], role)

        if url:
            domain = urlparse(url).netloc.replace("www.", "")
            link = ap.cell(ap_row, ap_headers["Job Link"], f"{domain} ↗")
            link.hyperlink = url
            link.font = Font(name="Arial", color="0563C1", underline="single")

        date_applied_col = ap_headers.get("Date Applied")
        if date_applied_col:
            dc = ap.cell(ap_row, date_applied_col, date.today())
            dc.number_format = "yyyy-mm-dd"

        ap.cell(ap_row, ap_headers["Location"], location)
        ap.cell(ap_row, ap_headers["Status"], status)
        if "Cover Letter" in ap_headers:
            ap.cell(ap_row, ap_headers["Cover Letter"], "No")
        if "Notes" in ap_headers:
            ap.cell(ap_row, ap_headers["Notes"], f"(Migrated from New Jobs) {fit_notes or ''}".strip())

        for c in range(1, ap.max_column + 1):
            cell = ap.cell(ap_row, c)
            if not (c == ap_headers.get("Job Link") and url):
                cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        migrated += 1

    if migrate_rows:
        # Clear all data rows (values + hyperlinks) then rewrite only the ones we keep,
        # in order, from row 2 — avoids delete_rows' hyperlink-relocation bug entirely.
        max_existing_row = nj.max_row
        for r_idx in range(2, max_existing_row + 1):
            for c_idx in range(1, n_cols + 1):
                cell = nj.cell(r_idx, c_idx)
                cell.value = None
                cell.hyperlink = None

        for i, r in enumerate(keep_rows):
            r_idx = 2 + i
            vals = r["values"]
            for c_idx in range(1, n_cols + 1):
                cell = nj.cell(r_idx, c_idx)
                cell.value = vals[c_idx - 1]
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c_idx == nj_headers.get("Date Found"):
                    cell.number_format = "yyyy-mm-dd"
            link_col = nj_headers["Job Link"]
            link_cell = nj.cell(r_idx, link_col)
            if r["url"]:
                link_cell.hyperlink = r["url"]
                link_cell.font = Font(name="Arial", color="0563C1", underline="single")

        last_row = 1 + len(keep_rows)
        if "Applications" in ap.tables:
            ap.tables["Applications"].ref = f"A1:{ap.cell(ap.max_row, ap.max_column).coordinate}"
        if "NewJobs" in nj.tables:
            nj.tables["NewJobs"].ref = f"A1:{nj.cell(max(last_row, 2), n_cols).coordinate}"

if applied or migrated:
    wb.save(XLSX)
print(f"applied {applied} web-edited row(s) to xlsx; migrated {migrated} row(s) from New Jobs to Applications")
