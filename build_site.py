#!/usr/bin/env python3
"""Generate data.json + index.html for the job tracker GitHub Pages site from Job_Application_Tracker.xlsx."""
import datetime
import json
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "Job_Application_Tracker.xlsx")
OUT_HTML = os.path.join(HERE, "index.html")
OUT_JSON = os.path.join(HERE, "data.json")

def sheet_rows(ws):
    rows = []
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        if all(c.value in (None, "") for c in row):
            continue
        rec = {}
        for h, c in zip(headers, row):
            val = c.value
            if isinstance(val, (datetime.date, datetime.datetime)):
                val = val.strftime("%Y-%m-%d")
            rec[h] = val
            if h in ("Job URL", "Job Link") and c.hyperlink:
                rec["_url"] = c.hyperlink.target
        if "example row" in str(rec.get("Role", "")).lower():
            continue
        rows.append(rec)
    return rows

wb = openpyxl.load_workbook(XLSX)
data = {
    "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
    "applications": sheet_rows(wb["Applications"]),
    "new_jobs": sheet_rows(wb["New Jobs"]) if "New Jobs" in wb.sheetnames else [],
}

with open(OUT_JSON, "w") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

page = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Jestin Roy — Job Tracker</title>
<style>
  :root {
    --bg: #f6f7f9; --card: #ffffff; --text: #1a1d21; --muted: #667085;
    --border: #e4e7ec; --accent: #1f4e78; --link: #0563c1;
  }
  @media (prefers-color-scheme: dark) {
    :root {
      --bg: #101418; --card: #1a2027; --text: #e6e9ec; --muted: #98a2b3;
      --border: #2c343d; --accent: #6ea8dc; --link: #7ab8f5;
    }
  }
  * { box-sizing: border-box; margin: 0; }
  body { background: var(--bg); color: var(--text); font: 15px/1.55 -apple-system, "Segoe UI", Roboto, Arial, sans-serif; padding: 2rem 1rem 4rem; }
  .wrap { max-width: 1100px; margin: 0 auto; }
  h1 { font-size: 1.6rem; margin-bottom: .25rem; }
  .sub { color: var(--muted); margin-bottom: 1rem; font-size: .9rem; }
  .bar { display: flex; gap: .6rem; flex-wrap: wrap; align-items: center; margin-bottom: 2rem; }
  .counts { display: flex; gap: .6rem; flex-wrap: wrap; }
  .counts span { background: var(--card); border: 1px solid var(--border); border-radius: 999px; padding: .3rem .85rem; font-size: .85rem; color: var(--muted); }
  .counts b { color: var(--text); }
  button { cursor: pointer; border: 1px solid var(--border); background: var(--card); color: var(--text); border-radius: 8px; padding: .35rem .9rem; font-size: .85rem; }
  button.primary { background: var(--accent); border-color: var(--accent); color: #fff; }
  button:disabled { opacity: .45; cursor: default; }
  h2 { font-size: 1.15rem; margin: 2rem 0 .75rem; }
  .tablewrap { overflow-x: auto; background: var(--card); border: 1px solid var(--border); border-radius: 12px; }
  table { border-collapse: collapse; width: 100%; min-width: 780px; }
  th { text-align: left; font-size: .75rem; text-transform: uppercase; letter-spacing: .05em; color: var(--muted); padding: .7rem .9rem; border-bottom: 1px solid var(--border); }
  td { padding: .7rem .9rem; border-bottom: 1px solid var(--border); vertical-align: top; }
  tr:last-child td { border-bottom: none; }
  a { color: var(--link); text-decoration: none; }
  a:hover { text-decoration: underline; }
  select { border: 1px solid var(--border); background: var(--card); color: var(--text); border-radius: 6px; padding: .15rem .3rem; font-size: .82rem; }
  [contenteditable="true"] { outline: 1px dashed var(--border); border-radius: 4px; padding: .1rem .25rem; min-width: 120px; display: inline-block; }
  .badge { display: inline-block; padding: .15rem .6rem; border-radius: 999px; font-size: .78rem; font-weight: 600; white-space: nowrap; }
  .s-applied, .s-submitted { background: #ddebf7; color: #1f4e78; }
  .s-in-progress { background: #fce4d6; color: #8a4b12; }
  .s-interview { background: #fff2cc; color: #7a6000; }
  .s-offer { background: #c6efce; color: #1d6b30; }
  .s-rejected, .s-not-proceeding { background: #e2e4e8; color: #555c66; }
  .s-saved { background: #e8e0f7; color: #4b3a80; }
  @media (prefers-color-scheme: dark) {
    .s-applied, .s-submitted { background: #1d3a57; color: #a8cdef; }
    .s-in-progress { background: #4a2f1a; color: #f0b98a; }
    .s-interview { background: #4a4213; color: #e8d67a; }
    .s-offer { background: #1d4426; color: #93d9a4; }
    .s-rejected, .s-not-proceeding { background: #333940; color: #9aa3ad; }
    .s-saved { background: #322a4d; color: #c3b3f0; }
  }
  .muted { color: var(--muted); }
  .titlebar { display: flex; align-items: center; gap: .7rem; }
  .social { display: flex; gap: .45rem; }
  .social a { display: inline-flex; align-items: center; justify-content: center; width: 30px; height: 30px; border-radius: 8px; border: 1px solid var(--border); background: var(--card); color: var(--muted); transition: color .15s, border-color .15s; }
  .social a:hover { color: var(--accent); border-color: var(--accent); text-decoration: none; }
  .social svg { width: 16px; height: 16px; fill: currentColor; }
  .notes { font-size: .85rem; color: var(--muted); max-width: 340px; display: inline-block; }
  #msg { font-size: .85rem; }
  footer { margin-top: 2.5rem; color: var(--muted); font-size: .8rem; }
</style>
</head>
<body>
<div class="wrap">
  <div class="titlebar">
    <h1>Job Application Tracker</h1>
    <nav class="social">
      <a href="https://www.linkedin.com/in/jestinroy3/" target="_blank" rel="noopener" aria-label="LinkedIn" title="LinkedIn">
        <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M20.45 20.45h-3.55v-5.57c0-1.33-.03-3.04-1.85-3.04-1.86 0-2.14 1.45-2.14 2.94v5.67H9.35V9h3.41v1.56h.05c.47-.9 1.63-1.85 3.36-1.85 3.6 0 4.27 2.37 4.27 5.46v6.28zM5.34 7.43a2.06 2.06 0 1 1 0-4.12 2.06 2.06 0 0 1 0 4.12zM7.12 20.45H3.56V9h3.56v11.45z"/></svg>
      </a>
      <a href="https://github.com/mr-jestin-roy" target="_blank" rel="noopener" aria-label="GitHub" title="GitHub">
        <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 .5A11.5 11.5 0 0 0 .5 12c0 5.08 3.29 9.39 7.86 10.91.58.11.79-.25.79-.55v-2.17c-3.2.7-3.87-1.36-3.87-1.36-.52-1.33-1.28-1.68-1.28-1.68-1.04-.71.08-.7.08-.7 1.15.08 1.76 1.18 1.76 1.18 1.03 1.76 2.69 1.25 3.34.96.1-.75.4-1.26.72-1.55-2.55-.29-5.24-1.28-5.24-5.68 0-1.26.45-2.28 1.18-3.09-.12-.29-.51-1.46.11-3.05 0 0 .97-.31 3.17 1.18a11 11 0 0 1 5.78 0c2.2-1.49 3.16-1.18 3.16-1.18.63 1.59.24 2.76.12 3.05.74.81 1.18 1.83 1.18 3.09 0 4.41-2.7 5.38-5.26 5.67.41.35.77 1.05.77 2.12v3.15c0 .3.2.67.8.55A11.5 11.5 0 0 0 23.5 12 11.5 11.5 0 0 0 12 .5z"/></svg>
      </a>
    </nav>
  </div>
  <div class="sub">Jestin Roy · synced with Job_Application_Tracker.xlsx · last build <span id="gen"></span></div>
  <div class="bar">
    <div class="counts" id="counts"></div>
    <span style="flex:1"></span>
    <button id="editBtn">✏️ Edit</button>
    <button id="saveBtn" class="primary" disabled>Save changes</button>
    <button id="tokenBtn" title="Set GitHub token">🔑</button>
    <span id="msg" class="muted"></span>
  </div>
  <h2>Applications</h2>
  <div class="tablewrap"><table id="apps"></table></div>
  <h2>New jobs found (last scans)</h2>
  <div class="tablewrap"><table id="newjobs"></table></div>
  <footer>Edits made here are committed to data.json and folded back into the xlsx on the next scheduled sync (12am / 12pm AEST).</footer>
</div>
<script>
const REPO = "mr-jestin-roy/job-tracker";
const API = `https://api.github.com/repos/${REPO}/contents/data.json`;
const STATUSES = ["Saved","In progress","Applied","Submitted","Interview","Offer","Rejected","Not proceeding"];
let DATA = null, editing = false, dirty = false;

const esc = v => (v == null ? "" : String(v).replace(/&/g,"&amp;").replace(/</g,"&lt;"));
const badge = s => s ? `<span class="badge s-${s.toLowerCase().replace(/\s+/g,"-")}">${esc(s)}</span>` : "";
const linkCell = r => {
  const url = r._url || r["Job URL"] || null;
  return url ? `<a href="${esc(url)}" target="_blank" rel="noopener">${new URL(url).hostname.replace("www.","")} ↗</a>` : "";
};
const statusCell = (r, list, i) => editing
  ? `<select data-list="${list}" data-i="${i}">` + STATUSES.map(s =>
      `<option ${s===r.Status?"selected":""}>${s}</option>`).join("") + "</select>"
  : badge(r.Status);
const notesCell = (r, key, list, i) =>
  `<span class="notes" ${editing?`contenteditable="true" data-list="${list}" data-i="${i}" data-key="${key}"`:""}>${esc(r[key])}</span>`;

function render() {
  document.getElementById("gen").textContent = DATA.generated;
  const mk = (id, rows, cols) => {
    const t = document.getElementById(id);
    t.innerHTML = "<thead><tr>" + cols.map(c=>`<th>${c[0]}</th>`).join("") + "</tr></thead><tbody>" +
      rows.map((r,i)=>"<tr>"+cols.map(c=>`<td>${c[1](r,i)}</td>`).join("")+"</tr>").join("") + "</tbody>";
  };
  mk("apps", DATA.applications, [
    ["Company",(r)=>`<b>${esc(r.Company)}</b>`],
    ["Role",(r)=>esc(r.Role)],
    ["Link",linkCell],
    ["Applied",(r)=>esc(r["Date Applied"])],
    ["Location",(r)=>esc(r.Location)],
    ["Status",(r,i)=>statusCell(r,"applications",i)],
    ["Notes",(r,i)=>notesCell(r,"Notes","applications",i)],
  ]);
  mk("newjobs", DATA.new_jobs, [
    ["Found",(r)=>esc(r["Date Found"])],
    ["Company",(r)=>`<b>${esc(r.Company)}</b>`],
    ["Role",(r)=>esc(r.Role)],
    ["Location",(r)=>esc(r.Location)],
    ["Link",linkCell],
    ["Posted",(r)=>esc(r.Posted)],
    ["Fit",(r,i)=>notesCell(r,"Fit Notes","new_jobs",i)],
    ["Status",(r,i)=>statusCell(r,"new_jobs",i)],
  ]);
  const a = DATA.applications;
  const active = a.filter(r=>!["Rejected","Not proceeding"].includes(r.Status)).length;
  document.getElementById("counts").innerHTML =
    `<span><b>${a.length}</b> applications</span><span><b>${active}</b> active</span>` +
    `<span><b>${a.filter(r=>r.Status==="Interview").length}</b> interviews</span>` +
    `<span><b>${DATA.new_jobs.length}</b> new jobs queued</span>`;
  bindEditors();
}

function markDirty(list, i) {
  DATA[list][i]._edited = new Date().toISOString();
  dirty = true;
  document.getElementById("saveBtn").disabled = false;
}

function bindEditors() {
  document.querySelectorAll("select[data-list]").forEach(el => el.onchange = () => {
    DATA[el.dataset.list][el.dataset.i].Status = el.value;
    markDirty(el.dataset.list, el.dataset.i);
  });
  document.querySelectorAll("[contenteditable][data-list]").forEach(el => el.oninput = () => {
    DATA[el.dataset.list][el.dataset.i][el.dataset.key] = el.innerText;
    markDirty(el.dataset.list, el.dataset.i);
  });
}

function getToken(force) {
  let t = localStorage.getItem("jt_token");
  if (!t || force) {
    t = prompt("Paste a fine-grained GitHub token with Contents read/write on " + REPO +
      ".\nStored only in this browser (localStorage).");
    if (t) localStorage.setItem("jt_token", t.trim());
  }
  return localStorage.getItem("jt_token");
}

async function save() {
  const token = getToken(false);
  const msg = document.getElementById("msg");
  if (!token) { msg.textContent = "No token set."; return; }
  msg.textContent = "Saving…";
  try {
    const headers = { Authorization: `Bearer ${token}`, Accept: "application/vnd.github+json" };
    const cur = await (await fetch(API, { headers })).json();
    const content = btoa(unescape(encodeURIComponent(JSON.stringify(DATA, null, 2))));
    const res = await fetch(API, {
      method: "PUT", headers,
      body: JSON.stringify({ message: "Web edit: update tracker data", content, sha: cur.sha }),
    });
    if (!res.ok) throw new Error((await res.json()).message || res.status);
    dirty = false;
    document.getElementById("saveBtn").disabled = true;
    msg.textContent = "Saved ✓ (xlsx syncs on next scheduled run)";
  } catch (e) {
    msg.textContent = "Save failed: " + e.message;
  }
}

document.getElementById("editBtn").onclick = () => {
  editing = !editing;
  document.getElementById("editBtn").textContent = editing ? "👁 View" : "✏️ Edit";
  render();
};
document.getElementById("saveBtn").onclick = save;
document.getElementById("tokenBtn").onclick = () => getToken(true);

fetch("data.json?t=" + Date.now()).then(r => r.json()).then(d => { DATA = d; render(); });
</script>
</body>
</html>
"""

with open(OUT_HTML, "w") as f:
    f.write(page)
print("wrote", OUT_JSON, "and", OUT_HTML)
